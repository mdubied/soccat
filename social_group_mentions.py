import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Step 1: Read the CSV file using the 'python' engine
csv_file = 'Social_Group_mentions.csv'
df = pd.read_csv(csv_file, engine='python')

# Step 2: Parse the 'HasGroup' column to extract the start and stop positions
def parse_hasgroup(hasgroup):
    if pd.isna(hasgroup):
        return []
    return eval(hasgroup)

df['ParsedHasGroup'] = df['HasGroup'].apply(parse_hasgroup)

# Step 3: Create a new Excel workbook and add the sentences with bolded group mentions
wb = Workbook()
ws = wb.active
ws.title = 'Highlighted Sentences'


# Function to apply bold to specific parts of a sentence
def apply_bold_to_sentence(ws, row_num, sentence, groups):
    cell = ws.cell(row=row_num, column=1)
    normal_font = Font(bold=False)
    bold_font = Font(bold=True)

    current_run = ''
    font = normal_font
    col_pos = 1
    pos = 0
    for start, stop, _ in groups:
        if pos < start:
            current_run = sentence[pos:start]
            ws.cell(row=row_num, column=col_pos).value = current_run
            ws.cell(row=row_num, column=col_pos).font = font
            pos = start
            col_pos = col_pos+1
        font = bold_font
        current_run = sentence[start:stop]
        ws.cell(row=row_num, column=col_pos).value = current_run
        ws.cell(row=row_num, column=col_pos).font = font
        pos = stop
        col_pos = col_pos+1
        font = normal_font
    if pos < len(sentence):
        current_run = sentence[pos:]
        ws.cell(row=row_num, column=col_pos).value = current_run
        ws.cell(row=row_num, column=col_pos).font = font

# Total number of rows
total_rows = len(df)-1
row_num = ws.max_row

# Add the sentences to the worksheet with bold formatting
for index, row in df.iterrows():
    print(index,'/',total_rows)
    
    sentence = row['sentence']
    groups = row['ParsedHasGroup']
    apply_bold_to_sentence(ws, row_num, sentence, groups)
    row_num = ws.max_row + 1

# Save the workbook
output_file = 'Social_Group_mentions_highlighted.xlsx'
wb.save(output_file)
